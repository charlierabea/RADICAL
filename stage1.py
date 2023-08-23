import glob
import os
import openpyxl
import torch
from PIL import Image
import open_clip

def get_classification_scores_for_study(study_dir, binary_tasks, template, model, preprocess_val, tokenizer, device):
    image_files = glob.glob(os.path.join(study_dir, '*'))
    images = torch.stack([preprocess_val(Image.open(img)) for img in image_files]).to(device)

    all_scores = {}

    for task in binary_tasks:
        condition, healthy = task, "healthy"
        labels = [template + condition, template + healthy]
        texts = tokenizer(labels, context_length=context_length).to(device)

        with torch.no_grad():
            image_features, text_features, logit_scale = model(images, texts)
            logits = (logit_scale * image_features @ text_features.t()).detach().softmax(dim=-1)
            logits = logits.cpu().numpy()
        
        for img_idx, img_name in enumerate(image_files):
            if img_name not in all_scores:
                all_scores[img_name] = {}
            all_scores[img_name][condition] = logits[img_idx, 0]

    return image_files, all_scores

def save_to_excel(data, filename):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write headers
    headers = ["Image Name"] + list(list(data.values())[0].keys())
    for col_num, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        sheet[f"{col_letter}1"] = header

    # Write data
    for row_num, (img_name, scores) in enumerate(data.items(), 2):
        sheet[f"A{row_num}"] = img_name
        for col_num, (condition, score) in enumerate(scores.items(), 2):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            sheet[f"{col_letter}{row_num}"] = score

    workbook.save(filename)

def main(patient_numbers, root_directory, binary_tasks, template):
    detailed_scores = {}
    average_scores = {}

    for patient in patient_numbers:
        patient_dir = os.path.join(root_directory, patient)
        
        # Loop through studies of each patient
        for study_dir in glob.glob(os.path.join(patient_dir, '*')):
            image_files, logits = get_classification_scores_for_study(study_dir, binary_tasks, template, model, preprocess_val, tokenizer, device)

            # For each image, record scores
            for img, scores in zip(image_files, logits.values()):
                img_name = os.path.basename(img)
                detailed_scores[f"{patient}_{os.path.basename(study_dir)}_{img_name}"] = scores

            # Compute average scores for the study
            study_name = os.path.basename(study_dir)
            scores_for_study = {img: detailed_scores[img] for img in detailed_scores if study_name in img}
            average_scores[f"{patient}_{study_name}"] = {condition: sum(scores_for_study[img][condition] for img in scores_for_study) / len(scores_for_study) for condition in binary_tasks}

    # Save detailed scores to Excel
    save_to_excel(detailed_scores, "/mnt/e/alzheimer_anonymized/packages//BiomedCLIP-PubMedBERT_256-vit_base_patch16_224/detailed_scores.xlsx")

    # Save average scores for each study to Excel
    save_to_excel(average_scores, "/mnt/e/alzheimer_anonymized/packages//BiomedCLIP-PubMedBERT_256-vit_base_patch16_224/average_scores.xlsx")


# Load the BiomedCLIP model and the preprocessing functions
model, preprocess_train, preprocess_val = open_clip.create_model_and_transforms('hf-hub:microsoft/BiomedCLIP-PubMedBERT_256-vit_base_patch16_224')
tokenizer = open_clip.get_tokenizer('hf-hub:microsoft/BiomedCLIP-PubMedBERT_256-vit_base_patch16_224')

# Define your binary classification tasks
binary_tasks = [
    'brain atrophy',
    'meningioma',
    'fracture',
    'soft tissue swelling',
    'hydrocephalus',
    'low density patches',
    'enlargement of the ventriclar system',
    'enlargement of the sulci',
    'calcified plaques',
    'midline shift',
    'intracranial hepatoma',
    'epidural hepatoma',
    'subdural hepatoma',
    'lacunar infarction',
    'cortical infarction',
    'subcortical infarction',
    'herniation',
    'arteriosclerotic Encephalopathy'
]

# Set up the device
device = torch.device('cuda') if torch.cuda.is_available() else torch.device('cpu')
model.to(device)
model.eval()

# Define root directory and patient numbers
root_directory = "/mnt/e/alzheimer_anonymized/alzheimer_2011-Glaucoma_anonymized/"
patient_numbers = ['A_1', 'A_2', 'A_3', 'A_4', 'A_5',
                   'A_6', 'A_7', 'A_8', 'A_9', 'A_10',
                   'A_11', 'A_12', 'A_13', 'A_14', 'A_15',
                   'A_16', 'A_17', 'A_18', 'A_19', 'A_20']

template = '{} presented in image'
context_length = 256

# Call the main function
main(patient_numbers, root_directory, binary_tasks, template)
